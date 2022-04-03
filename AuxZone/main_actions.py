import os
from datetime import date
from datetime import datetime
import mysql.connector

import openpyxl
import pandas as pd
import unidecode
from PyQt5.QtWidgets import QMessageBox, QFileDialog
from openpyxl.styles import Alignment, Font
from pandas import DataFrame

import main_panels
import main_toolbar


class C_MainActions():
    def __init__(self):

        ################ Pegando instancias definidas no Main

        self.MainWindowToolBar = main_toolbar.C_MenuToolBar()
        self.initUI()

    def initUI(self):  ### Instanciando os objetos
        self.MainPainel = main_panels.C_MainPanel(self)

    def execAbout(self):
        pass

    # Método executado para selecionar arquivo para captação dos endereços
    def execopenFileNamesDialog(self):
        self.MainPainel.Main_seleciona_endereco_QPushButton.setEnabled(False)
        self.MainWindowToolBar.Net_Select_Act.setEnabled(False)

        """
        Inicialmente esse métódo irá testar se existe algum arquivo na área de arrasta e solta,
        caso não, ele irá abrir uma janela Dialog com usuário para que seja feira a seleção.
        """

        # Variavel que armazena infomação que esta na aea de arrasta e solta do software.
        arquivo = self.MainPainel.Main_Recebe_Arquivo_QPlainTextEdit.toPlainText()

        # Criando uma variável adcional que recebe o objeto que a variável arquivo recebeu
        # Como o objeto é um caminho C:// é interessante obter apenas o nome do arquivo sem o prefixo
        # A variavel abaixo utiliza o método basename da biblioteca OS pra realizar essa operação.
        nome_do_arquivo = os.path.basename(arquivo)

        """
        Iremos realizar um teste para verificar se na variável arquivo existe alguma informação, ou seja,
        se foi utilizado a opção de arrastar e soltar para captação do arquivo.
        """

        if arquivo != "":  # Se a variável arquivo for diferente de vázio.

            """
            Só chegamos nessa parte do código pois foi utilizada a opção arrastar e soltar.
            Nesse caso iremos modificar a estrutura da string desse objeto e armazenar a string modificada
            numa outra variável, que chamaremos de arquivo_de_referencia
            """

            arquivo_de_referencia = arquivo.replace("file:///", "")

            """
            Vamos instanciar três métodos diferentes: um para selecionar os endereços dos clientes
                                                      um para identificar se existem clientes vips
                                                      um para mostrar a lista de endereços e clientes vips
            Os métodos serão instanciados a baixo passando como parâmetro: o caminho do arquivo selecionado.
                                                                           e no ultimo o nome do arquivo que será 
                                                                           utilizado para ser mostrado ao usuário
            """

            self.nome_arquivo = self.nome_do_arquivo(nome_do_arquivo.replace(".txt", ""))
            try:
                data_atual = date.today()
                os.mkdir("Zoneamento_" + str(data_atual))
                self.execSalvar_arquivo_em_excel(arquivo_de_referencia, nome_do_arquivo)
            except:
                self.execSalvar_arquivo_em_excel(arquivo_de_referencia, nome_do_arquivo)

        else:
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            arquivo_de_referencia, _ = QFileDialog.getOpenFileNames(None, "Selecione o Arquivo", "", "Files (*.txt)",
                                                                    options=options)

            if arquivo_de_referencia == []:
                pass
            else:

                nome_do_arquivo = os.path.basename(arquivo_de_referencia[0])

                # Esse procedimento é importante para retirar a primeira linha do arquivo aberto.
                with open(arquivo_de_referencia[0], "r+") as f:
                    new_f = f.readlines()
                    f.seek(0)
                    for line in new_f:
                        if "#Contratos#########" not in line:
                            f.write(line)
                    f.truncate()
                    f.close()

                if arquivo_de_referencia:
                    self.MainPainel.Main_Recebe_Arquivo_QPlainTextEdit.setPlainText(nome_do_arquivo)
                    self.nome_arquivo = self.nome_do_arquivo(nome_do_arquivo.replace(".txt", ""))
                try:
                    data_atual = date.today()
                    os.mkdir("Zoneamento_" + str(data_atual))
                    self.execSalvar_arquivo_em_excel(arquivo_de_referencia, nome_do_arquivo)
                except:
                    self.execSalvar_arquivo_em_excel(arquivo_de_referencia, nome_do_arquivo)

    def execSalvar_arquivo_em_excel(self, arquivo_de_referencia, nome_do_arquivo):
        nome_do_arquivo = nome_do_arquivo.replace(".txt", "")
        df_arquivo_de_referencia = pd.read_table(arquivo_de_referencia[0], sep='#')
        df_arquivo_de_referencia.to_excel("Zoneamento_" + str(date.today()) + "/" + nome_do_arquivo + ".xlsx",
                                          sheet_name=nome_do_arquivo, na_rep="#N/A", header=True, index=False)

        self.get_Endereco(df_arquivo_de_referencia)
        self.get_Enderecos_vips(df_arquivo_de_referencia)

    def exec_Convert_arquivo_unico_para_excel(self):

        try:
            data_atual = date.today()
            os.mkdir("Conversoes_" + str(data_atual))
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            arquivo_de_referencia_que_sera_convertido, _ = QFileDialog.getOpenFileNames(None, "Selecione o Arquivo", "",
                                                                                        "Files (*.txt)",
                                                                                        options=options)
            if arquivo_de_referencia_que_sera_convertido == []:
                pass
            else:
                nome_do_arquivo_que_sera_covertido_para_excel = os.path.basename(
                    arquivo_de_referencia_que_sera_convertido[0])
                # Esse procedimento é importante para retirar a primeira linha do arquivo aberto.
                with open(arquivo_de_referencia_que_sera_convertido[0], "r+") as f:
                    new_f = f.readlines()
                    f.seek(0)
                    for line in new_f:
                        if "#Contratos#########" not in line:
                            f.write(line)
                    f.truncate()
                    f.close()
                nome_do_arquivo_que_sera_covertido_para_excel = nome_do_arquivo_que_sera_covertido_para_excel.replace(
                    ".txt", "")
                df_arquivo_de_referencia_que_sera_convertido_para_excel = pd.read_table(
                    arquivo_de_referencia_que_sera_convertido[0], sep='#')
                df_arquivo_de_referencia_que_sera_convertido_para_excel.to_excel(
                    "Conversoes_" + str(date.today()) + "/" + nome_do_arquivo_que_sera_covertido_para_excel + ".xlsx",
                    sheet_name=nome_do_arquivo_que_sera_covertido_para_excel, na_rep="#N/A", header=True, index=False)
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setWindowTitle("Conversor")
                msg.setText("Arquivo convertido e salvo na pasta " + "Conversoes_" + str(date.today()))
                msg.exec()
                # QMessageBox(self, "Title" , "Message")
        except:
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            arquivo_de_referencia_que_sera_convertido, _ = QFileDialog.getOpenFileNames(None, "Selecione o Arquivo", "",
                                                                                        "Files (*.txt)",
                                                                                        options=options)
            if arquivo_de_referencia_que_sera_convertido == []:
                pass
            else:
                nome_do_arquivo_que_sera_covertido_para_excel = os.path.basename(
                    arquivo_de_referencia_que_sera_convertido[0])
                # Esse procedimento é importante para retirar a primeira linha do arquivo aberto.
                with open(arquivo_de_referencia_que_sera_convertido[0], "r+") as f:
                    new_f = f.readlines()
                    f.seek(0)
                    for line in new_f:
                        if "#Contratos#########" not in line:
                            f.write(line)
                    f.truncate()
                    f.close()
                nome_do_arquivo_que_sera_covertido_para_excel = nome_do_arquivo_que_sera_covertido_para_excel.replace(
                    ".txt", "")
                df_arquivo_de_referencia_que_sera_convertido_para_excel = pd.read_table(
                    arquivo_de_referencia_que_sera_convertido[0], sep='#')
                df_arquivo_de_referencia_que_sera_convertido_para_excel.to_excel(
                    "Conversoes_" + str(date.today()) + "/" + nome_do_arquivo_que_sera_covertido_para_excel + ".xlsx",
                    sheet_name=nome_do_arquivo_que_sera_covertido_para_excel, na_rep="#N/A", header=True, index=False)
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setWindowTitle("Conversor")
                msg.setText("Arquivo convertido e salvo na pasta " + "Conversoes_" + str(date.today()))
                msg.exec()

    def exec_Convert_varios_arquivos_de_uma_pasta(self):
        try:
            data_atual = date.today()
            os.mkdir("Conversoes_" + str(data_atual))
            self.dir_path = QFileDialog.getExistingDirectory(None, "Selecione a Pasta", "")
            for _, _, arquivo_de_referencia_que_sera_convertido in os.walk(self.dir_path):
                if self.dir_path == "":
                    pass
                else:

                    if ".txt" in arquivo_de_referencia_que_sera_convertido[0]:
                        print(arquivo_de_referencia_que_sera_convertido[0])
                        for i in range(0, len(arquivo_de_referencia_que_sera_convertido)):
                            nome_do_arquivo_que_sera_covertido_para_excel = arquivo_de_referencia_que_sera_convertido[i]
                            # Esse procedimento é importante para retirar a primeira linha do arquivo aberto.
                            with open(
                                    (self.dir_path + "/" + nome_do_arquivo_que_sera_covertido_para_excel, "r+")) as f:
                                new_f = f.readlines()
                                f.seek(0)
                                for line in new_f:
                                    if "#Contratos#########" not in line:
                                        f.write(line)
                                f.truncate()
                                f.close()
                            df_arquivo_de_referencia_que_sera_convertido_para_excel = pd.read_table(
                                self.dir_path + "/" + nome_do_arquivo_que_sera_covertido_para_excel,
                                                    sep='#')
                            nome_do_arquivo_que_sera_covertido_para_excel = nome_do_arquivo_que_sera_covertido_para_excel.replace(
                                ".txt", "")
                            df_arquivo_de_referencia_que_sera_convertido_para_excel.to_excel(
                                "Conversoes_" + str(
                                    date.today()) + "/" + nome_do_arquivo_que_sera_covertido_para_excel + ".xlsx",
                                                    sheet_name=nome_do_arquivo_que_sera_covertido_para_excel,
                                                    na_rep="#N/A", header=True,
                                                    index=False)
                        msg = QMessageBox()
                        msg.setIcon(QMessageBox.Information)
                        msg.setWindowTitle("Conversor")
                        msg.setText("Arquivos convertidos e salvos na pasta " + "Conversoes_" + str(date.today()))
                        msg.exec()
                    else:
                        msg = QMessageBox()
                        msg.setIcon(QMessageBox.Warning)
                        msg.setWindowTitle("Error")
                        msg.setText("A pasta selecionada não possui os arquivos corretos")
                        msg.exec()

        except:

            self.dir_path = QFileDialog.getExistingDirectory(None, "Selecione a Pasta", "")
            for _, _, arquivo_de_referencia_que_sera_convertido in os.walk(self.dir_path):
                if self.dir_path == "":
                    pass
                else:

                    if ".txt" in arquivo_de_referencia_que_sera_convertido[0]:
                        for i in range(0, len(arquivo_de_referencia_que_sera_convertido)):
                            nome_do_arquivo_que_sera_covertido_para_excel = arquivo_de_referencia_que_sera_convertido[i]
                            # Esse procedimento é importante para retirar a primeira linha do arquivo aberto.
                            with open(self.dir_path + "/" + nome_do_arquivo_que_sera_covertido_para_excel, "r+") as f:
                                new_f = f.readlines()
                                f.seek(0)
                                for line in new_f:
                                    if "#Contratos#########" not in line:
                                        f.write(line)
                                f.truncate()
                                f.close()
                            df_arquivo_de_referencia_que_sera_convertido_para_excel = pd.read_table(
                                self.dir_path + "/" + nome_do_arquivo_que_sera_covertido_para_excel, sep='#')
                            nome_do_arquivo_que_sera_covertido_para_excel = nome_do_arquivo_que_sera_covertido_para_excel.replace(
                                ".txt", "")
                            df_arquivo_de_referencia_que_sera_convertido_para_excel.to_excel("Conversoes_" + str(
                                date.today()) + "/" + nome_do_arquivo_que_sera_covertido_para_excel + ".xlsx",
                                                                                             sheet_name=nome_do_arquivo_que_sera_covertido_para_excel,
                                                                                             na_rep="#N/A", header=True,
                                                                                             index=False)
                        msg = QMessageBox()
                        msg.setIcon(QMessageBox.Information)
                        msg.setWindowTitle("Conversor")
                        msg.setText("Arquivos convertidos e salvos na pasta " + "Conversoes_" + str(date.today()))
                        msg.exec()
                    else:
                        msg = QMessageBox()
                        msg.setIcon(QMessageBox.Warning)
                        msg.setWindowTitle("Error")
                        msg.setText("A pasta selecionada não possui os arquivos corretos")
                        msg.exec()

    def nome_do_arquivo(self, nome_do_arquivo):
        return nome_do_arquivo

    def get_Endereco(self, df_arquivo_de_referencia):
        numero_de_linhas = df_arquivo_de_referencia[df_arquivo_de_referencia.columns[0]].count()
        lista_de_enderecos = []
        lista_de_enderecos_filtradas_sem_virgula = []
        for i in range(0, numero_de_linhas):
            lista_de_enderecos.append(df_arquivo_de_referencia.loc[i, "Endereco"])
        lista_de_enderecos_filtradas = sorted(set(lista_de_enderecos))
        sep = ','
        for i in range(0, len(lista_de_enderecos_filtradas)):
            res = lista_de_enderecos_filtradas[i].split(sep, 1)[0]
            lista_de_enderecos_filtradas_sem_virgula.append(res)
        lista_de_enderecos_filtradas_sem_virgula = sorted(set(lista_de_enderecos_filtradas_sem_virgula))

        tamanho_da_lista = len(lista_de_enderecos_filtradas_sem_virgula)
        for i in range(0, tamanho_da_lista):
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("PO ",
                                                                                                              "Povoado ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("ASSENT ",
                                                                                                              "Assentamento ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace(
                "ASSENAMENTO", "Assentamento")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("FZ ",
                                                                                                              "Fazenda ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("ET ",
                                                                                                              "Estrada ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("AV ",
                                                                                                              "Av.")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("LO ",
                                                                                                              "Loteamento ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("TV ",
                                                                                                              "Travessa ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("PC ",
                                                                                                              "Praça ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("RO ",
                                                                                                              "Rodovia ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("SI ",
                                                                                                              "Sitio ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("CJ ",
                                                                                                              "Conjunto ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("Cj ",
                                                                                                              "Conjunto ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("Lg ",
                                                                                                              "Largo ")
            
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].title()
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("Dos ",
                                                                                                              "dos ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("Das ",
                                                                                                              "das ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("Do ",
                                                                                                              "do ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("Da ",
                                                                                                              "da ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("De ",
                                                                                                              "de ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("Ii",
                                                                                                              "II")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("Iii",
                                                                                                              "III")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("Vi ",
                                                                                                              "VI")
        self.join_list_de_enderecos(lista_de_enderecos_filtradas_sem_virgula)

    def join_list_de_enderecos(self, lista_de_enderecos_filtradas_sem_virgula):
        if len(lista_de_enderecos_filtradas_sem_virgula) == 1:
            Enderecos = '{}.'.format(
                lista_de_enderecos_filtradas_sem_virgula[0])  # adicionar o ultimo com 'e' e retornar
            self.execPrintEndereco(Enderecos)
        else:
            str = ', '.join(lista_de_enderecos_filtradas_sem_virgula[
                            :-1])  # gerar uma string com os items separados por virgula, com excecao do ultimo
            Enderecos = '{} e {}.'.format(str, lista_de_enderecos_filtradas_sem_virgula[
                -1])  # adicionar o ultimo com 'e' e retornar
            self.execPrintEndereco(Enderecos)

    def get_Endereco_zone(self, df_arquivo_de_referencia):
        numero_de_linhas = df_arquivo_de_referencia[df_arquivo_de_referencia.columns[0]].count()
        lista_de_enderecos = []
        lista_de_enderecos_filtradas_sem_virgula = []
        for i in range(0, numero_de_linhas):
            lista_de_enderecos.append(df_arquivo_de_referencia.loc[i, "Endereco"])
        lista_de_enderecos_filtradas = sorted(set(lista_de_enderecos))
        sep = ','
        for i in range(0, len(lista_de_enderecos_filtradas)):
            res = lista_de_enderecos_filtradas[i].split(sep, 1)[0]
            lista_de_enderecos_filtradas_sem_virgula.append(res)
        lista_de_enderecos_filtradas_sem_virgula = sorted(set(lista_de_enderecos_filtradas_sem_virgula))

        tamanho_da_lista = len(lista_de_enderecos_filtradas_sem_virgula)
        for i in range(0, tamanho_da_lista):
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("PO ",
                                                                                                              "Povoado ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("ASSENT ",
                                                                                                              "Assentamento ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace(
                "ASSENAMENTO", "Assentamento")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("FZ ",
                                                                                                              "Fazenda ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("ET ",
                                                                                                              "Estrada ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("AV ",
                                                                                                              "Av.")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("LO ",
                                                                                                              "Loteamento ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("TV ",
                                                                                                              "Travessa ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("PC ",
                                                                                                              "Praça ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("Ro ",
                                                                                                              "Rodovia ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("SI ",
                                                                                                              "Sitio ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("CJ ",
                                                                                                              "Conjunto ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("Cj ",
                                                                                                              "Conjunto ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("Lg ",
                                                                                                              "Largo ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].title()
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("Dos ",
                                                                                                              "dos ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("Das ",
                                                                                                              "das ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("Do ",
                                                                                                              "do ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("Da ",
                                                                                                              "da ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("De ",
                                                                                                              "de ")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("Ii",
                                                                                                              "II")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("Iii",
                                                                                                              "III")
            lista_de_enderecos_filtradas_sem_virgula[i] = lista_de_enderecos_filtradas_sem_virgula[i].replace("Vi ",
                                                                                                              "VI")
        self.join_list_de_enderecos_zone(lista_de_enderecos_filtradas_sem_virgula)

    def join_list_de_enderecos_zone(self, lista_de_enderecos_filtradas_sem_virgula):
        if len(lista_de_enderecos_filtradas_sem_virgula) == 1:
            self.Enderecos_para_o_zoneamento = '{}.'.format(
                lista_de_enderecos_filtradas_sem_virgula[0])  # adicionar o ultimo com 'e' e retornar
            return self.Enderecos_para_o_zoneamento
        else:
            str = ', '.join(lista_de_enderecos_filtradas_sem_virgula[
                            :-1])  # gerar uma string com os items separados por virgula, com excecao do ultimo
            self.Enderecos_para_o_zoneamento = '{} e {}.'.format(str, lista_de_enderecos_filtradas_sem_virgula[
                -1])  # adicionar o ultimo com 'e' e retornar
            return self.Enderecos_para_o_zoneamento

    def get_Enderecos_vips(self, df_arquivo_de_referencia):
        numero_de_linhas = df_arquivo_de_referencia[df_arquivo_de_referencia.columns[0]].count()
        lista_de_enderecos_vips = {}
        for i in range(0, numero_de_linhas):
            tipos_de_vips = ["Vip 1", "Vip 2", "Vip 3", "Vip 4", "Vip 5", "Vip 6", "Vip 7", "Vip 8", "Vip 9", "Vip 10"]
            if df_arquivo_de_referencia.loc[i, "Vip"] in tipos_de_vips:
                lista_de_enderecos_vips[str(df_arquivo_de_referencia.loc[i, "Nome"])] = str(
                    df_arquivo_de_referencia.loc[i, "Vip"])
                if df_arquivo_de_referencia.loc[i, "Vip"] == ("Vip 1" or "Vip 10"):
                    msg = QMessageBox()
                    msg.setIcon(QMessageBox.Warning)
                    msg.setWindowTitle("Atenção")
                    msg.setText("Nesse Desligamento Existe um Cliente VIP 1 ou 10")
                    msg.exec()

        self.execPrintEnderecoVip(lista_de_enderecos_vips)

    def exec_adcionar_informacoes_na_si(self):

        lista_de_sis_do_zoneamento = []
        self.diretorio_dos_arquivos_a_serem_enviados = QFileDialog.getExistingDirectory(None, "Selecione a Pasta", "")
        for _, _, arquivos_a_serem_enviados in os.walk(self.diretorio_dos_arquivos_a_serem_enviados):
            if arquivos_a_serem_enviados != "":
                options = QFileDialog.Options()
                options |= QFileDialog.DontUseNativeDialog
                arquivo_de_base_do_zoneamento, _ = QFileDialog.getOpenFileNames(None,
                                                                                "Selecione o Arquivo Base do Zoneamento Atual",
                                                                                "", "Files (*.xlsx)", options=options)
                wb_conjunto_de_si = openpyxl.load_workbook(filename=str(arquivo_de_base_do_zoneamento[0]))

                for i in range(0, len(arquivos_a_serem_enviados)):
                    SI_para_pesquisa = arquivos_a_serem_enviados[i].replace(".xlsx", "")
                    lista_de_sis_do_zoneamento.append(SI_para_pesquisa)

        for i in range(0, len(lista_de_sis_do_zoneamento)):
            wb_si_individual = openpyxl.load_workbook(filename=str(
                self.diretorio_dos_arquivos_a_serem_enviados + "/" + lista_de_sis_do_zoneamento[i] + ".xlsx"))
            ws_informacoes = wb_si_individual.create_sheet("Informações SI")
            ws_informacoes.column_dimensions['A'].width = 20
            ws_informacoes.column_dimensions['B'].width = 30
            ws_informacoes['A1'] = "INÍCIO"
            ws_informacoes['A1'].alignment = Alignment(horizontal='left')
            ws_informacoes['A1'].font = Font(bold=True)
            ws_informacoes['A2'] = "HORA INICIAL"
            ws_informacoes['A2'].alignment = Alignment(horizontal='left')
            ws_informacoes['A2'].font = Font(bold=True)
            ws_informacoes['A3'] = "TÉRMINO"
            ws_informacoes['A3'].alignment = Alignment(horizontal='left')
            ws_informacoes['A3'].font = Font(bold=True)
            ws_informacoes['A4'] = "HORA FINAL"
            ws_informacoes['A4'].alignment = Alignment(horizontal='left')
            ws_informacoes['A4'].font = Font(bold=True)
            ws_informacoes['A5'] = "BAIRRO"
            ws_informacoes['A5'].alignment = Alignment(horizontal='left')
            ws_informacoes['A5'].font = Font(bold=True)
            ws_conjunto_de_si = wb_conjunto_de_si.worksheets[0]
            for j in range(1, ws_conjunto_de_si.max_row + 1):
                wb_conjunto_de_si = openpyxl.load_workbook(filename=str(arquivo_de_base_do_zoneamento[0]))
                ws_conjunto_de_si = wb_conjunto_de_si.worksheets[0]
                if (lista_de_sis_do_zoneamento[i] + "/2021") == ws_conjunto_de_si["E" + str(j)].value:
                    wb_si_individual.active
                    ws_informacoes['B1'] = ws_conjunto_de_si["A" + str(j)].value
                    ws_informacoes['B1'].alignment = Alignment(horizontal='left')
                    ws_informacoes['B2'] = ws_conjunto_de_si["B" + str(j)].value
                    ws_informacoes['B2'].alignment = Alignment(horizontal='left')
                    ws_informacoes['B3'] = ws_conjunto_de_si["C" + str(j)].value
                    ws_informacoes['B3'].alignment = Alignment(horizontal='left')
                    ws_informacoes['B4'] = ws_conjunto_de_si["D" + str(j)].value
                    ws_informacoes['B4'].alignment = Alignment(horizontal='left')
                    ws_informacoes['B5'] = ws_conjunto_de_si["AV" + str(j)].value
                    ws_informacoes['B5'].alignment = Alignment(horizontal='left')
            wb_si_individual.save(
                str(self.diretorio_dos_arquivos_a_serem_enviados + "/" + lista_de_sis_do_zoneamento[i] + ".xlsx"))
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle("Atualização")
        msg.setText("Os arquivos da pasta Conversoes_ " + str(
            date.today()) + "\n" + "Foram atualizados com as informações da" + "\n" + "Solicitação de Intervenção")
        msg.exec()

    def exec_gerar_arquivo_final_zoneamento(self):
        data_atual = date.today()

        self.lista_de_arquivos_que_foram_gerados_para_o_zoneamento =[]

        DIAS = [
            'Segunda-feira',
            'Terca-feira',
            'Quarta-feira',
            'Quinta-feira',
            'Sexta-feira',
            'Sabado',
            'Domingo'
        ]

        self.lista_data_segunda = []
        self.lista_data_terca = []
        self.lista_data_quarta = []
        self.lista_data_quinta = []
        self.lista_data_sexta = []
        self.lista_data_sabado = []
        self.lista_data_domingo = []

        lista_de_sis_do_zoneamento = []
        self.diretorio_dos_arquivos_que_criaram_relatorio = QFileDialog.getExistingDirectory(None,"Selecione a Pasta que contém os arquivos excel atualizados", "")
        for _, _, arquivos_do_relatorio in os.walk(self.diretorio_dos_arquivos_que_criaram_relatorio):
            if arquivos_do_relatorio != "":

                for i in range(0, len(arquivos_do_relatorio)):
                    lista_de_sis_do_zoneamento.append(arquivos_do_relatorio[i])
        try:
            os.remove("Arquivo_Zoneamento_Segunda-feira_" + str(
                    data_atual) + "/" + "Arquivo_Zoneamento_Segunda-feira_" + str(data_atual) + ".txt")
            os.remove("Arquivo_Zoneamento_Terca-feira_" + str(
                data_atual) + "/" + "Arquivo_Zoneamento_Terca-feira_" + str(data_atual) + ".txt")
            os.remove("Arquivo_Zoneamento_Quarta-feira_" + str(
                data_atual) + "/" + "Arquivo_Zoneamento_Quarta-feira_" + str(data_atual) + ".txt")
            os.remove("Arquivo_Zoneamento_Quinta-feira_" + str(
                data_atual) + "/" + "Arquivo_Zoneamento_Quinta-feira_" + str(data_atual) + ".txt")
            os.remove("Arquivo_Zoneamento_Sexta-feira_" + str(
                data_atual) + "/" + "Arquivo_Zoneamento_Sexta-feira_" + str(data_atual) + ".txt")
            os.remove("Arquivo_Zoneamento_Sabado_" + str(
                data_atual) + "/" + "Arquivo_Zoneamento_Sabado_" + str(data_atual) + ".txt")
            os.remove("Arquivo_Zoneamento_Domingo_" + str(
                data_atual) + "/" + "Arquivo_Zoneamento_Domingo_" + str(data_atual) + ".txt")
        except:
            pass

        print(lista_de_sis_do_zoneamento)

        for i in range(0, len(lista_de_sis_do_zoneamento)):
            data_frame_final_zone = pd.read_excel(
                self.diretorio_dos_arquivos_que_criaram_relatorio + "/" + arquivos_do_relatorio[i])
            self.get_Endereco_zone(data_frame_final_zone)
            enderecamento_da_si = self.Enderecos_para_o_zoneamento
            wb_si_do_zoneamento = openpyxl.load_workbook(
                filename=str(self.diretorio_dos_arquivos_que_criaram_relatorio + "/" + arquivos_do_relatorio[i]))
            ws_informacao_zoneamento = wb_si_do_zoneamento.worksheets[1]
            dia_da_semana = ws_informacao_zoneamento['B1'].value
            municipio_bairro = ws_informacao_zoneamento['B5'].value
            hora_inicial = ws_informacao_zoneamento['B2'].value
            hora_inicial = str(hora_inicial)[:-3]
            hora_final = ws_informacao_zoneamento['B4'].value
            hora_final = str(hora_final)[:-3]
            numero_da_si = arquivos_do_relatorio[i].replace(".xlsx", "")
            data = dia_da_semana
            data_do_dia_da_semana = dia_da_semana
            indice_da_semana = data.weekday()
            dia_da_semana = DIAS[indice_da_semana]

            if dia_da_semana == 'Segunda-feira':
                data_modificada = str(data).replace("-", "/")
                d = data_modificada[:-9]
                data_em_texto = '{}/{}/{}'.format(d[8:10], d[5:7], d[0:4])
                self.lista_data_segunda.append(data_em_texto)
                data_atual = date.today()

                if not os.path.exists("Arquivo_Zoneamento_Segunda-feira_" + str(data_atual)):
                    diretorio = os.mkdir("Arquivo_Zoneamento_Segunda-feira_" + str(data_atual))
                arquivo_s = open("Arquivo_Zoneamento_Segunda-feira_" + str(
                    data_atual) + "/" + "Arquivo_Zoneamento_Segunda-feira_" + str(data_atual) + ".txt", "a")
                arquivo_s.write(
                    str(municipio_bairro) + " - das " + str(hora_inicial) + " às " + str(hora_final) + " - " + str(
                        enderecamento_da_si) + "\n" + "SI " + str(numero_da_si) + "/2021" + 3 * "\n")
                arquivo_s.close()

            if dia_da_semana == 'Terca-feira':
                data_modificada = str(data).replace("-", "/")
                d = data_modificada[:-9]
                data_em_texto = '{}/{}/{}'.format(d[8:10], d[5:7], d[0:4])
                self.lista_data_terca.append(data_em_texto)
                data_atual = date.today()
                if not os.path.exists("Arquivo_Zoneamento_Terca-feira_" + str(data_atual)):
                    diretorio = os.mkdir("Arquivo_Zoneamento_Terca-feira_" + str(data_atual))
                arquivo_t = open(
                    "Arquivo_Zoneamento_Terca-feira_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Terca-feira_" + str(
                        data_atual) + ".txt", "a")
                arquivo_t.write(
                    str(municipio_bairro) + " - das " + str(hora_inicial) + " às " + str(hora_final) + " - " + str(
                        enderecamento_da_si) + "\n" + "SI " + str(numero_da_si) + "/2021" + 3 * "\n")
                arquivo_t.close()

            if dia_da_semana == 'Quarta-feira':
                data_modificada = str(data).replace("-", "/")
                d = data_modificada[:-9]
                data_em_texto = '{}/{}/{}'.format(d[8:10], d[5:7], d[0:4])
                self.lista_data_quarta.append(data_em_texto)
                data_atual = date.today()
                if not os.path.exists("Arquivo_Zoneamento_Quarta-feira_" + str(data_atual)):
                    diretorio = os.mkdir("Arquivo_Zoneamento_Quarta-feira_" + str(data_atual))
                arquivo_qua = open("Arquivo_Zoneamento_Quarta-feira_" + str(
                    data_atual) + "/" + "Arquivo_Zoneamento_Quarta-feira_" + str(data_atual) + ".txt", "a")
                arquivo_qua.write(
                    str(municipio_bairro) + " - das " + str(hora_inicial) + " às " + str(hora_final) + " - " + str(
                        enderecamento_da_si) + "\n" + "SI " + str(numero_da_si) + "/2021" + 3 * "\n")
                arquivo_qua.close()


            if dia_da_semana == 'Quinta-feira':
                data_modificada = str(data).replace("-", "/")
                d = data_modificada[:-9]
                data_em_texto = '{}/{}/{}'.format(d[8:10], d[5:7], d[0:4])
                self.lista_data_quinta.append(data_em_texto)
                data_atual = date.today()
                if not os.path.exists("Arquivo_Zoneamento_Quinta-feira_" + str(data_atual)):
                    diretorio = os.mkdir("Arquivo_Zoneamento_Quinta-feira_" + str(data_atual))
                arquivo_qui = open("Arquivo_Zoneamento_Quinta-feira_" + str(
                    data_atual) + "/" + "Arquivo_Zoneamento_Quinta-feira_" + str(data_atual) + ".txt", "a")
                arquivo_qui.write(
                    str(municipio_bairro) + " - das " + str(hora_inicial) + " às " + str(hora_final) + " - " + str(
                        enderecamento_da_si) + "\n" + "SI " + str(numero_da_si) + "/2021" + 3 * "\n")
                arquivo_qui.close()

            if dia_da_semana == 'Sexta-feira':
                data_modificada = str(data).replace("-", "/")
                d = data_modificada[:-9]
                data_em_texto = '{}/{}/{}'.format(d[8:10], d[5:7], d[0:4])
                self.lista_data_sexta.append(data_em_texto)
                data_atual = date.today()
                if not os.path.exists("Arquivo_Zoneamento_Sexta-feira_" + str(data_atual)):
                    diretorio = os.mkdir("Arquivo_Zoneamento_Sexta-feira_" + str(data_atual))
                arquivo_sex = open(
                    "Arquivo_Zoneamento_Sexta-feira_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Sexta-feira_" + str(
                        data_atual) + ".txt", "a")
                arquivo_sex.write(
                    str(municipio_bairro) + " - das " + str(hora_inicial) + " às " + str(hora_final) + " - " + str(
                        enderecamento_da_si) + "\n" + "SI " + str(numero_da_si) + "/2021" + 3 * "\n")
                arquivo_sex.close()

            if dia_da_semana == 'Sabado':
                data_modificada = str(data).replace("-", "/")
                d = data_modificada[:-9]
                data_em_texto = '{}/{}/{}'.format(d[8:10], d[5:7], d[0:4])
                self.lista_data_sabado.append(data_em_texto)
                data_atual = date.today()
                if not os.path.exists("Arquivo_Zoneamento_Sabado_" + str(data_atual)):
                    diretorio = os.mkdir("Arquivo_Zoneamento_Sabado_" + str(data_atual))
                arquivo_sab = open(
                    "Arquivo_Zoneamento_Sabado_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Sabado_" + str(
                        data_atual) + ".txt", "a")
                arquivo_sab.write(
                    str(municipio_bairro) + " - das " + str(hora_inicial) + " às " + str(hora_final) + " - " + str(
                        enderecamento_da_si) + "\n" + "SI " + str(numero_da_si) + "/2021" + 3 * "\n")
                arquivo_sab.close()

            if dia_da_semana == 'Domingo':
                data_modificada = str(data).replace("-", "/")
                d = data_modificada[:-9]
                data_em_texto = '{}/{}/{}'.format(d[8:10], d[5:7], d[0:4])
                self.lista_data_domingo.append(data_em_texto)
                data_atual = date.today()
                if not os.path.exists("Arquivo_Zoneamento_Domingo_" + str(data_atual)):
                    diretorio = os.mkdir("Arquivo_Zoneamento_Domingo_" + str(data_atual))
                arquivo_dom = open(
                    "Arquivo_Zoneamento_Domingo_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Domingo_" + str(
                        data_atual) + ".txt", "a")
                arquivo_dom.write(
                    str(municipio_bairro) + " - das " + str(hora_inicial) + " às " + str(hora_final) + " - " + str(
                        enderecamento_da_si) + "\n" + "SI " + str(numero_da_si) + "/2021" + 3 * "\n")
                arquivo_dom.close()

        try:
            data_atual = date.today()
            self.lista_de_arquivos_que_foram_gerados_para_o_zoneamento.append("Arquivo_Zoneamento_Segunda-feira_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Segunda-feira_" + str(
                            data_atual) + ".txt")
            with open(
                    "Arquivo_Zoneamento_Segunda-feira_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Segunda-feira_" + str(
                            data_atual) + ".txt", 'r') as arquivo:
                linhas = arquivo.readlines()
            with open(
                    "Arquivo_Zoneamento_Segunda-feira_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Segunda-feira_" + str(
                            data_atual) + ".txt", 'w') as arquivo:
                arquivo.writelines(self.lista_data_segunda[0] + " SEGUNDA-FEIRA " + 3 * '\n')
                arquivo.writelines(linhas)
            arquivo.close()
        except:
            pass

        try:
            data_atual = date.today()
            self.lista_de_arquivos_que_foram_gerados_para_o_zoneamento.append("Arquivo_Zoneamento_Terca-feira_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Terca-feira_" + str(
                    data_atual) + ".txt")
            with open("Arquivo_Zoneamento_Terca-feira_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Terca-feira_" + str(
                    data_atual) + ".txt", 'r') as arquivo:
                linhas = arquivo.readlines()
            with open("Arquivo_Zoneamento_Terca-feira_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Terca-feira_" + str(
                    data_atual) + ".txt", 'w') as arquivo:
                arquivo.writelines(self.lista_data_terca[0] + " TERCA-FEIRA " + 3 * '\n')
                arquivo.writelines(linhas)
            arquivo.close()
        except:
            pass

        try:
            data_atual = date.today()
            self.lista_de_arquivos_que_foram_gerados_para_o_zoneamento.append("Arquivo_Zoneamento_Quarta-feira_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Quarta-feira_" + str(
                    data_atual) + ".txt")
            with open("Arquivo_Zoneamento_Quarta-feira_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Quarta-feira_" + str(
                    data_atual) + ".txt", 'r') as arquivo:
                linhas = arquivo.readlines()
            with open("Arquivo_Zoneamento_Quarta-feira_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Quarta-feira_" + str(
                    data_atual) + ".txt", 'w') as arquivo:
                arquivo.writelines(self.lista_data_quarta[0] + " QUARTA-FEIRA " + 3 * '\n')
                arquivo.writelines(linhas)
            arquivo.close()
        except:
            pass

        try:
            data_atual = date.today()
            self.lista_de_arquivos_que_foram_gerados_para_o_zoneamento.append("Arquivo_Zoneamento_Quinta-feira_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Quinta-feira_" + str(
                    data_atual) + ".txt")
            with open("Arquivo_Zoneamento_Quinta-feira_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Quinta-feira_" + str(
                    data_atual) + ".txt", 'r') as arquivo:
                linhas = arquivo.readlines()
            with open("Arquivo_Zoneamento_Quinta-feira_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Quinta-feira_" + str(
                    data_atual) + ".txt", 'w') as arquivo:
                arquivo.writelines(self.lista_data_quinta[0] + " QUINTA-FEIRA " + 3 * '\n')
                arquivo.writelines(linhas)
            arquivo.close()
        except:
            pass


        try:
            data_atual = date.today()
            self.lista_de_arquivos_que_foram_gerados_para_o_zoneamento.append("Arquivo_Zoneamento_Sexta-feira_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Sexta-feira_" + str(
                    data_atual) + ".txt")
            with open("Arquivo_Zoneamento_Sexta-feira_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Sexta-feira_" + str(
                    data_atual) + ".txt", 'r') as arquivo:
                linhas = arquivo.readlines()
            with open("Arquivo_Zoneamento_Sexta-feira_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Sexta-feira_" + str(
                    data_atual) + ".txt", 'w') as arquivo:
                arquivo.writelines(self.lista_data_sexta[0] + " SEXTA-FEIRA " + 3 * '\n')
                arquivo.writelines(linhas)
            arquivo.close()
        except:
            pass

        try:
            data_atual = date.today()
            self.lista_de_arquivos_que_foram_gerados_para_o_zoneamento.append("Arquivo_Zoneamento_Sabado_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Sabado_" + str(
                    data_atual) + ".txt")
            with open("Arquivo_Zoneamento_Sabado_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Sabado_" + str(
                    data_atual) + ".txt", 'r') as arquivo:
                linhas = arquivo.readlines()
            with open("Arquivo_Zoneamento_Sabado_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Sabado_" + str(
                    data_atual) + ".txt", 'w') as arquivo:
                arquivo.writelines(self.lista_data_sabado[0] + " SÁBADO " + 3 * '\n')
                arquivo.writelines(linhas)
            arquivo.close()
        except:
            pass

        try:
            data_atual = date.today()
            self.lista_de_arquivos_que_foram_gerados_para_o_zoneamento.append("Arquivo_Zoneamento_Domingo_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Domingo_" + str(
                    data_atual) + ".txt")
            with open(
                    "Arquivo_Zoneamento_Domingo_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Domingo_" + str(
                            data_atual) + ".txt", 'r') as arquivo:
                linhas = arquivo.readlines()
            with open(
                    "Arquivo_Zoneamento_Domingo_" + str(data_atual) + "/" + "Arquivo_Zoneamento_Domingo_" + str(
                            data_atual) + ".txt", 'w') as arquivo:
                arquivo.writelines(self.lista_data_domingo[0] + " DOMINGO " + 3 * '\n')
                arquivo.writelines(linhas)
            arquivo.close()
        except:
            pass

        try:
            with open("Zoneamento.txt", "w") as file:
                print(self.lista_de_arquivos_que_foram_gerados_para_o_zoneamento)

                # Percorre a lista de arquivos a serem lidos:
                for i in range(0,len(self.lista_de_arquivos_que_foram_gerados_para_o_zoneamento)):
                    for temp in [self.lista_de_arquivos_que_foram_gerados_para_o_zoneamento[i]]:
                        # Abre cada arquivo para leitura:
                        with open(temp, "r") as t:
                            # Escreve no arquivo o conteúdo:
                            file.writelines("\n")
                            file.writelines(t)
        except:
            pass


    def exec_gerar_arquivo_final_zoneamento_teste1(self):
        print("aqui")

        DIAS = [
            'Segunda-feira',
            'Terça-feira',
            'Quarta-feira',
            'Quinta-Feira',
            'Sexta-feira',
            'Sábado',
            'Domingo'
        ]

        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        arquivo_final_do_zoneamento, _ = QFileDialog.getOpenFileNames(None,
                                                                      "Selecione o Arquivo Base do Zoneamento Atual",
                                                                      "", "Files (*.xlsx)", options=options)
        wb_final_zone = openpyxl.load_workbook(filename=str(arquivo_final_do_zoneamento[0]))
        data_frame_final_zone = pd.read_excel(arquivo_final_do_zoneamento[0])
        data_frame_final_zone_ordenado = pd.DataFrame(data_frame_final_zone.sort_values(by="INÍCIO"))
        data_atual = date.today()
        if not os.path.exists("Arquivo_Zoneamento_" + str(data_atual)):
            diretorio = os.mkdir("Arquivo_Zoneamento_" + str(data_atual))
        arquivo = open("Arquivo_Zoneamento_" + str(data_atual) + "/" + "Arquivo_Zoneamento_" + str(data_atual) + ".txt",
                       "a")

        for i in range(0, data_frame_final_zone_ordenado[data_frame_final_zone_ordenado.columns[0]].count()):
            data = data_frame_final_zone_ordenado.loc[i, "INÍCIO"]
            inicio = data_frame_final_zone_ordenado.loc[i, "HORA INICIAL"]
            fim = data_frame_final_zone_ordenado.loc[i, "HORA FINAL"]
            data = data_frame_final_zone_ordenado.loc[i, "INÍCIO"]
            indice_da_semana = data.weekday()
            dia_da_semana = DIAS[indice_da_semana]
            if dia_da_semana == "Quarta-feira":
                arquivo.seek(0)
                arquivo.write(str(data) + " " + "QUARTA-FEIRA" + "\n")
                arquivo.write("das " + str(inicio)[0:5] )








    def exec_Enviar_Email(self):
        self.diretorio_dos_arquivos_a_serem_enviados = QFileDialog.getExistingDirectory(None, "Selecione a Pasta", "")
        for _, _, arquivos_a_serem_enviados in os.walk(self.diretorio_dos_arquivos_a_serem_enviados):
            if arquivos_a_serem_enviados != "":
                options = QFileDialog.Options()
                options |= QFileDialog.DontUseNativeDialog
                arquivo_de_base_do_zoneamento, _ = QFileDialog.getOpenFileNames(None,
                                                                                "Selecione o Arquivo Base do Zoneamento Atual",
                                                                                "", "Files (*.xlsx)",
                                                                                options=options)
                wb = openpyxl.load_workbook(filename=str(arquivo_de_base_do_zoneamento[0]))
                ws = wb.active
                df = DataFrame(ws.values)

                for i in range(0, len(arquivos_a_serem_enviados)):
                    print(arquivos_a_serem_enviados[i])

    def execPrintEndereco(self, Enderecos):
        self.MainPainel.Main_Recebe_Endereco_QPlainTextEdit.setPlainText(Enderecos)

    def execPrintEnderecoVip(self, Enderecos_vips):
        self.MainPainel.escreve_enderecos_vips(Enderecos_vips)

    def exec_limpar_area_de_arquivo(self):
        self.MainPainel.Main_Recebe_Arquivo_QPlainTextEdit.clear()
        self.MainPainel.Main_Recebe_Endereco_QPlainTextEdit.clear()
        self.MainPainel.Main_Recebe_Endereco_vip_QPlainTextEdit.clear()
        self.MainPainel.Main_seleciona_endereco_QPushButton.setEnabled(True)
        self.MainWindowToolBar.Net_Select_Act.setEnabled(True)

    def conectar_banco(self):
        print("1")
        con = mysql.connector.connect(host="ADM-2K12-161", user='REMOTO', password='COELBA')#db= "REMOTO@ADM-2K12-161"
        print("2")
        if con.is_connected():
            db_info = con.get_server_info()
            print("Conectado ao servidor MySQL versão ", db_info)

